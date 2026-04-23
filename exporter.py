"""
exporter.py — Export collected business data to styled Excel (.xlsx).

Output: <location>_business_data.xlsx with 4 sheets:
  Restaurants | Lodges | Homestay_Resorts | Dhabas
"""

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import logging

logger = logging.getLogger(__name__)

SHEET_CONFIG = {
    'Restaurants':      {'color': 'C0392B', 'icon': 'RESTAURANTS'},
    'Lodges':           {'color': '2980B9', 'icon': 'LODGES'},
    'Homestay_Resorts': {'color': '27AE60', 'icon': 'HOMESTAY & RESORTS'},
    'Dhabas':           {'color': 'E67E22', 'icon': 'DHABAS'},
}
COLUMNS = ['S.No', 'Name', 'Address', 'Phone', 'Alternate Phone']


def entries_to_df(entries: list, location: str, target: int = 75) -> pd.DataFrame:
    rows = []
    for i, e in enumerate(entries[:target], 1):
        rows.append({
            'S.No': i,
            'Name': e.get('name', '').strip(),
            'Address': e.get('address', location).strip(),
            'Phone': e.get('phone', ''),
            'Alternate Phone': e.get('alternate_phone', 'Not Available'),
        })
    while len(rows) < target:
        rows.append({
            'S.No': len(rows) + 1,
            'Name': 'Data not available',
            'Address': location,
            'Phone': 'N/A',
            'Alternate Phone': 'Not Available',
        })
    return pd.DataFrame(rows[:target], columns=COLUMNS)


def style_sheet(ws, sheet_name: str, location: str):
    cfg = SHEET_CONFIG.get(sheet_name, {'color': '2C3E50', 'icon': sheet_name})
    hdr_fill = PatternFill(start_color=cfg['color'], end_color=cfg['color'], fill_type='solid')
    hdr_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    alt_fill = PatternFill(start_color='EBF5FB', end_color='EBF5FB', fill_type='solid')
    title_fill = PatternFill(start_color='1A252F', end_color='1A252F', fill_type='solid')
    thin = Border(
        left=Side(style='thin', color='BDC3C7'),
        right=Side(style='thin', color='BDC3C7'),
        top=Side(style='thin', color='BDC3C7'),
        bottom=Side(style='thin', color='BDC3C7'),
    )

    # Insert 2 title rows
    ws.insert_rows(1)
    ws.insert_rows(1)

    ws['A1'] = f"{cfg['icon']} — {location.upper()} BUSINESS DIRECTORY"
    ws['A1'].font = Font(name='Calibri', bold=True, size=13, color='FFFFFF')
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f'A1:{get_column_letter(len(COLUMNS))}1')

    ws['A2'] = f'Location: {location.title()}   |   Total Entries: {ws.max_row - 3}   |   Category: {sheet_name.replace("_", " ")}'
    ws['A2'].font = Font(name='Calibri', italic=True, size=9, color='7F8C8D')
    ws['A2'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[2].height = 16
    ws.merge_cells(f'A2:{get_column_letter(len(COLUMNS))}2')

    # Header row (row 3)
    for col_num, col_name in enumerate(COLUMNS, 1):
        cell = ws.cell(row=3, column=col_num)
        cell.value = col_name
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin
    ws.row_dimensions[3].height = 20

    # Data rows
    for row in ws.iter_rows(min_row=4, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin
            cell.font = Font(name='Calibri', size=10)
            if row[0].row % 2 == 0:
                cell.fill = alt_fill

    # Column widths
    for i, w in enumerate([6, 38, 55, 16, 18], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A4'
    ws.auto_filter.ref = f'A3:{get_column_letter(len(COLUMNS))}{ws.max_row}'
    ws.sheet_view.zoomScale = 100


def save_temp_csv(entries: list, category: str, location: str, output_dir: str = '.'):
    if not entries:
        return
    os.makedirs(output_dir, exist_ok=True)
    path = os.path.join(output_dir, f'{location.lower()}_{category.lower()}_backup.csv')
    pd.DataFrame(entries).to_csv(path, index=False, encoding='utf-8-sig')
    logger.info(f'Backup CSV: {path}')


def export_to_excel(data: dict, location: str, output_dir: str = '.') -> str:
    os.makedirs(output_dir, exist_ok=True)
    filename = f"{location.lower().replace(' ', '_')}_business_data.xlsx"
    filepath = os.path.join(output_dir, filename)

    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        for sheet_name, entries in data.items():
            df = entries_to_df(entries, location)
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    wb = load_workbook(filepath)
    for sheet_name in data:
        if sheet_name in wb.sheetnames:
            style_sheet(wb[sheet_name], sheet_name, location)
    wb.save(filepath)
    logger.info(f'Excel saved: {filepath}')
    return filepath
